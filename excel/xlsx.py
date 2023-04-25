import openpyxl
from translate import translate


class Xlsx:

    def __init__(self, source):
        self.input_file = source
        self.out_file = source[:-5]+"_new.xlsx"
        self.workbook = openpyxl.load_workbook(source)
        self.translator = translate.Translator()

    def translate(self):
        workbook = self.workbook
        for worksheet_name in workbook.sheetnames:
            worksheet = workbook[worksheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    source = cell.value
                    try:
                        int(source)
                    except:
                        print(f"translating:{cell}")
                        translate = self.translator.translate(
                            source, sourceLang="zh",
                            toLang="en",
                        )
                        cell.value = translate
        workbook.save(self.out_file)
        print('Done')
