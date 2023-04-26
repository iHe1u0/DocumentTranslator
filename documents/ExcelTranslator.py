import tkinter.messagebox
import openpyxl

import translate
from translate import BaiduTranslator
from utils import file_utils


class ExcelTranslator:
    def __init__(self, source, out):
        if not file_utils.check_file_exist(source):
            tkinter.messagebox.showerror(title='提示', message='文件不存在')
        else:
            self.input_file = source
            self.out_file = out
            self.workbook = openpyxl.load_workbook(source)
            self.translator = BaiduTranslator.BaiduTranslator(app_id='', key='')

    def translate(self, source_lang='auto', des_lang='en'):
        workbook = self.workbook
        for worksheet_name in workbook.sheetnames:
            worksheet = workbook[worksheet_name]
            for row in worksheet.iter_rows():
                for cell in row:
                    source = cell.value
                    translate_result = self.translator.translate(
                        source,
                        source_lang=source_lang,
                        des_lang=des_lang,
                    )
                    cell.value = translate_result
        workbook.save(self.out_file)
